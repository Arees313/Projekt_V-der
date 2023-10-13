import openpyxl
from openpyxl.styles import Font
from openpyxl.styles import alignment
import requests
import json
import datetime
import pandas as pd
import sys
import time
import pandas as pd
now = datetime.datetime.now()
now_formaterad = now.strftime('%Y-%m-%d')
def huvud_meny():
    while True:
        välj = {
            '1': 'Hämta senaste data',
            '2': 'Skriv ut prognos',
            '9': 'Avsluta'

        }
        for key,value in välj.items():
            print(f"{key}. {value}")
        välj_input = input("Välj mellan 1,2 eller 9: ")
        if välj_input not in välj:
            print("Du måste välja mellan 1,2 och 9. ")
            time.sleep(1)
            print("Försök igen!")
            time.sleep(2)
            continue
        else:
            pass
        if välj_input == '1':
            huvudmeny_hämta_1()
        elif välj_input == '2':
            print("Skriver ut prognos..")
            time.sleep(1)
            print(f"Prognos från SMHI {now_formaterad}:")
            data_ur_excel = pd.read_excel('Väder_Samlaren.xlsx')
            for index, row in data_ur_excel.iterrows():
                timme = row['Hour']
                temp = row['Temperature']
                nederbörd = row['Rain or Snow']
                timme = f"{timme:02}:00"
                print(f'{timme} {temp} {nederbörd}')
            time.sleep(1)
            print("Återgår till menyn..")
            time.sleep(1)
            continue
        else:
            print("Programmet avslutas..")
            sys.exit()

def huvudmeny_hämta_1():
    while True:
        time.sleep(1)
        välj_2 = {
            '1': 'Hämta data från SMHI',
            '2': 'Hämta data från Openweathermap',
            '3': 'Hämta data från både SMHI och OWM',
            '4': 'Återgå till huvudmeny',
            '9': 'Avsluta'

        }

        for key,value in välj_2.items():
            print(f"{key}. {value}")
        välj_input_2 = input("Välj mellan 1,2,3 eller 9: ")
        
        if välj_input_2 == '9':
            print("Programmet avslutas..")
            sys.exit()     
        elif välj_input_2 == '1':
            url = (f'https://opendata-download-metfcst.smhi.se/api/category/pmp3g/version/2/geotype/point/lon/18.021515/lat/59.30996/data.json')
            def hämta_data_smhi():
                return requests.get(url)
            def SMHI_duplicerat(smhi:requests):
                smhi = smhi.json()
                viktig_data = smhi['timeSeries']
                temp_data = []
                for validtime in viktig_data[:25]:
                    validtime['parameters']
                    for value in validtime['parameters']:
                        if value['name'] == 't':
                            temp_data.append((value['values'][0]))
                return temp_data
            cash_data = SMHI_duplicerat(hämta_data_smhi())
            # cash_data = hämta_data_smhi().text
            # cash_data = hämta_data_smhi().json()
            # rich.print_json(data = cash_data)
            time.sleep(3)
            new_data = SMHI_duplicerat(hämta_data_smhi())
            #for old,new in zip(cash_data,new_data):
                #print(old==new, old, new)
            # new_data = get_important_data(hämta_data_smhi)
            if new_data == cash_data:
                print("Du har hämtat duplicerat data!") 
                time.sleep(1)
                print("Excel filen skapas, utan någon uppdaterad data..")
                inhämta_nyhet_smhi()
                time.sleep(1)
                print("Excel filen klar!")
                time.sleep(1)
                print("Återgår nu till menyn..")
                time.sleep(1)
                huvud_meny()
            else:
                print("Excel filen skapas, ny data implementeras..")
                inhämta_nyhet_smhi()
                time.sleep(1)
                print("Excel filen klar!")
                time.sleep(1)
                print("Återgår nu till menyn..")
                time.sleep(1)
                huvud_meny()
        elif välj_input_2 == '2':
            pass
        elif välj_input_2 == '3':
            pass
        elif välj_input_2 == '4':
            print("Återgår till huvudmenyn..")
            time.sleep(1.5)
            huvud_meny()



def inhämta_nyhet_smhi():
    now = datetime.datetime.now()
    now_formaterad = now.strftime('%Y-%m-%d')
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    nu = datetime.datetime.now()
    now = datetime.datetime.now()
    now_24 = now.strftime('%Y-%m-%d')
    now_24 = now + datetime.timedelta(hours=24)
    latitude = 59.30996552541549
    longitude = 18.02151508449004
    url = (f'https://opendata-download-metfcst.smhi.se/api/category/pmp3g/version/2/geotype/point/lon/18.021515/lat/59.30996/data.json')
    response = requests.get(url)
    data = json.loads(response.text)
    #Tänk nu på att API datan är från UTC 0 så om klockan är 13:00 i Sverige, är klockan 11:00 i UTC 0.
    category = 0
    bold = Font(bold=True)
    time_series_json = data['timeSeries'][:24]
    for time_series in time_series_json:
        for parameter in time_series['parameters']:           
            if parameter['unit'] == 'Cel':
                temp = parameter['values'][0]
            if parameter['name'] == 'pcat':
                nederbörd = parameter['values'][0]
        rad = 2
        while (now < now_24):
            nu111 = now + datetime.timedelta(hours=1)
            now_formaterad = now.strftime('%Y-%m-%d')
            now_hour_formaterad = nu111.strftime('%H')
            now = now + datetime.timedelta(hours=1)
            if nederbörd >= 1:
                parameter['name'] == 'pcat' and category <= 1
                result = True
            else:
                result = False
            temp_x = str(temp) + ' C'
            sheet.cell(row=rad, column = 1, value=nu)
            sheet.cell(row=rad, column = 2, value=longitude)
            sheet.cell(row=rad, column = 3, value=latitude)
            sheet.cell(row = rad,  column = 4, value=now_formaterad)
            sheet.cell(row = rad,  column = 5, value=now_hour_formaterad)
            sheet.cell(row = rad,  column = 6, value=temp_x)
            sheet.cell(row = rad,  column = 7, value=result)
            rad += 1
            now + datetime.timedelta(hours=1)
            break
    rad = sheet.cell(row=1, column=1, value='Created')
    rad.font = bold
    rad2 = sheet.cell(row=1, column=2, value='Longitude')
    rad2.font = bold
    rad3 = sheet.cell(row=1, column=3, value='Latitude')
    rad3.font = bold
    rad4 = sheet.cell(row=1, column=4, value='Date')
    rad4.font = bold
    rad5 = sheet.cell(row=1, column=5, value='Hour')
    rad5.font = bold
    rad6 = sheet.cell(row=1, column=6, value='Temperature')
    rad6.font = bold
    rad7 = sheet.cell(row=1, column=7, value='Rain or Snow')
    rad7.font = bold
    rad8 = sheet.cell(row=1, column=8, value='Provider')
    rad8.font = bold
    for row in range(2,26):
        sheet.cell(row=row, column=8, value='SMHI')
    workbook.save('Väder_Test.xlsx')
    workbook.close()


huvud_meny()